#encoding=utf-8
"""
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
# 创建文件对象
test_xl=Workbook()
# # 使用 active 获取第一个 sheet
test_sheet=test_xl.active
# # 写入列标题内容
test_sheet['a1']="列测试标题1"
test_sheet['b1']="列测试标题2"
test_sheet['c1']="列测试标题3"
# # 行写入
test_sheet.append({'A' : 10001, 'B' : "测试任务", "C" : datetime.datetime.now()})
# # 保存文件
test_xl.save("C:\\soft\\pycharm1\\test_xl\\demo2.xlsx")
# Workbook().get_sheet_names()：
# 以 list 格式返回 excel 中所有工作表名。
wb=load_workbook("C:\\soft\\pycharm1\\test_xl\\demo2.xlsx")
sheetnames=wb.get_sheet_names
print (sheetnames)
sheetnames1=wb.sheetnames
"""
from selenium import webdriver
from Base_paskage import open_browser
class Screenshot:
    def __init__(self):
        self.driver=open.browsers("chrom")






