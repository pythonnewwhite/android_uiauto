#coding=utf-8
import os
import openpyxl
import sys

base_path=os.getcwd()
print (base_path)
#打开表格
open_excel=openpyxl.load_workbook(base_path+"/Appium_Mj/config/case_login.xlsx")
#获取表格的sheet页
sheet_names=open_excel.sheetnames
#获取行数
sheet_name=open_excel[sheet_names[0]]
print(sheet_name)
#获取表格内容
value1=sheet_name.cell(1,2).value
print (value1)
#获取表格行数
sheet_lengths=sheet_name.max_row
print (sheet_name.max_row)

class Hand_Excel:
    def load_excel(self):
        #加载Excel表格
        open_excel=openpyxl.load_workbook(os.getcwd()+"/Appium_Mj/configg/tets.xlex")
        return open_excel

    def get_sheet_name(self,index):
        #获取需要的sheet页，根据index参数取对应的sheet表
        sheet_name=self.load_excel()[sheet_names[index]]
        return sheet_name

    def get_sheet_values(self,index,row,cols):
        #获取sheet页里面的具体内容,根据行数row、列数cols确定需要的值
        sheet_value=self.get_sheet_name(index).call_value(row,cols)
        return sheet_value

    def get_rows(self):
        pass



